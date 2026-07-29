import csv
import os
import tempfile

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors as reportlab_colors

from settings import get_company

NAVY = colors.HexColor("#1a237e")
LIGHT_NAVY = colors.HexColor("#e8eaf6")
BORDER = colors.HexColor("#cccccc")
WHITE = colors.white


def _get_font_name() -> str:
    from printer import FONT_NAME as printer_font
    return printer_font


FONT_NAME = _get_font_name()


def export_csv(rows: list[dict], column_headers: list[str], column_keys: list[str],
               output_path: str | None = None) -> str:
    if output_path is None:
        output_path = os.path.join(tempfile.gettempdir(), "report.csv")

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(column_headers)
        for row in rows:
            writer.writerow([row.get(k, "") for k in column_keys])
    return output_path


def export_xlsx(rows: list[dict], column_headers: list[str], column_keys: list[str],
                sheet_name: str = "Report",
                output_path: str | None = None) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if output_path is None:
        output_path = os.path.join(tempfile.gettempdir(), "report.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(column_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(column_keys, 1):
            val = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left")

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    wb.save(output_path)
    return output_path


def export_pdf_table(rows: list[dict], column_headers: list[str], column_keys: list[str],
                     title: str = "Report",
                     output_path: str | None = None) -> str:
    if output_path is None:
        output_path = os.path.join(tempfile.gettempdir(), "report.pdf")

    company = get_company()
    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=10*mm, rightMargin=10*mm,
        topMargin=12*mm, bottomMargin=12*mm,
    )

    styles = getSampleStyleSheet()
    s_title = ParagraphStyle("RptTitle", fontName=FONT_NAME, fontSize=16,
                             alignment=TA_CENTER, textColor=NAVY, spaceAfter=4)
    s_sub = ParagraphStyle("RptSub", fontName=FONT_NAME, fontSize=8,
                           alignment=TA_CENTER, textColor=colors.HexColor("#757575"),
                           spaceAfter=6*mm)
    s_hdr = ParagraphStyle("RptHdr", fontName=FONT_NAME, fontSize=7,
                           alignment=TA_CENTER, textColor=WHITE)
    s_cell = ParagraphStyle("RptCell", fontName=FONT_NAME, fontSize=7,
                            alignment=TA_CENTER)

    elements = []

    name = company.get("name", "").upper()
    if name:
        elements.append(Paragraph(name, s_title))
    elements.append(Paragraph(title, ParagraphStyle(
        "Title2", fontName=FONT_NAME, fontSize=12,
        alignment=TA_CENTER, textColor=NAVY, spaceAfter=4*mm,
    )))

    col_widths = [max(15*mm, 170*mm / len(column_headers))] * len(column_headers)
    data = [[Paragraph(h, s_hdr) for h in column_headers]]
    for row in rows:
        data.append([Paragraph(str(row.get(k, "")), s_cell) for k in column_keys])

    tbl = Table(data, colWidths=col_widths, repeatRows=1)

    tbl_style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, -1), FONT_NAME),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]
    if len(data) > 2:
        last = data[-1]
        if any(kw in str(last[0]) for kw in ["Total", "Grand", "Summary"]):
            tbl_style_cmds.append(("BACKGROUND", (0, -1), (-1, -1), LIGHT_NAVY))
            tbl_style_cmds.append(("LINEABOVE", (0, -1), (-1, -1), 0.8, NAVY))

    tbl.setStyle(TableStyle(tbl_style_cmds))

    if len(data) > 25:
        tbl_style_cmds.append(("FONTSIZE", (0, 1), (-1, -1), 6))

    elements.append(tbl)
    elements.append(Spacer(1, 4*mm))

    # Bank details (like bill print)
    bank_parts = []
    if company.get("bank_name"):
        bank_parts.append(f"Bank: {company['bank_name']}")
    if company.get("bank_account"):
        bank_parts.append(f"A/C: {company['bank_account']}")
    if company.get("bank_ifsc"):
        bank_parts.append(f"IFSC: {company['bank_ifsc']}")
    if bank_parts:
        bank_text = "&nbsp;&nbsp;|&nbsp;&nbsp;".join(bank_parts)
        bk_tbl = Table(
            [[Paragraph("<b>Bank Details:</b>",
                         ParagraphStyle("BKL", fontName=FONT_NAME, fontSize=8, textColor=colors.HexColor("#757575"))),
              Paragraph(bank_text,
                         ParagraphStyle("BKV", fontName=FONT_NAME, fontSize=8, textColor=colors.black))]],
            colWidths=[26*mm, 146*mm],
        )
        bk_tbl.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ]))
        elements.append(bk_tbl)
        elements.append(Spacer(1, 2*mm))

    from reportlab.platypus import HRFlowable
    elements.append(HRFlowable(width="100%", thickness=0.3, color=BORDER, spaceAfter=2*mm))
    elements.append(Paragraph(
        "This is a computer-generated report.",
        ParagraphStyle("Footer", fontName=FONT_NAME, fontSize=7,
                       alignment=TA_CENTER, textColor=colors.HexColor("#757575")),
    ))

    doc.build(elements)
    return output_path


def export_sales_register_pdf(
    bills: list[dict],
    receipts: list[dict],
    total_bills: float,
    total_receipts: float,
    balance: float,
    title: str = "Sales Register",
    date_from: str = "",
    date_to: str = "",
    output_path: str | None = None,
) -> str:
    if output_path is None:
        output_path = os.path.join(tempfile.gettempdir(), "sales_register.pdf")

    company = get_company()
    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=10*mm, rightMargin=10*mm,
        topMargin=12*mm, bottomMargin=12*mm,
    )

    styles = getSampleStyleSheet()
    s_title = ParagraphStyle("RptTitle", fontName=FONT_NAME, fontSize=16,
                             alignment=TA_CENTER, textColor=NAVY, spaceAfter=2)
    s_sub = ParagraphStyle("RptSub", fontName=FONT_NAME, fontSize=8,
                           alignment=TA_CENTER, textColor=colors.HexColor("#757575"),
                           spaceAfter=4*mm)
    s_hdr = ParagraphStyle("RptHdr", fontName=FONT_NAME, fontSize=7,
                           alignment=TA_CENTER, textColor=WHITE)
    s_cell = ParagraphStyle("RptCell", fontName=FONT_NAME, fontSize=7,
                            alignment=TA_CENTER)
    s_cell_left = ParagraphStyle("RptCellL", fontName=FONT_NAME, fontSize=7,
                                 alignment=TA_LEFT)
    s_section = ParagraphStyle("Section", fontName=FONT_NAME, fontSize=10,
                               textColor=NAVY, spaceBefore=4*mm, spaceAfter=2*mm)
    s_balance = ParagraphStyle("Balance", fontName=FONT_NAME, fontSize=12,
                               alignment=TA_RIGHT, textColor=NAVY, spaceBefore=2*mm,
                               spaceAfter=2*mm)

    elements = []

    name = company.get("name", "").upper()
    if name:
        elements.append(Paragraph(name, s_title))
    period_text = f"{title}"
    if date_from and date_to:
        period_text += f"  |  {date_from}  to  {date_to}"
    elements.append(Paragraph(period_text, s_sub))

    # --- Bills Table ---
    bill_hdr = ["#", "Date", "Bill No", "Customer", "Vehicle", "Gross", "Tare", "Net Wt", "Amount"]
    bill_cw = [6*mm, 18*mm, 26*mm, 30*mm, 20*mm, 14*mm, 14*mm, 16*mm, 20*mm]

    bill_data = [[Paragraph(h, s_hdr) for h in bill_hdr]]
    for i, b in enumerate(bills, 1):
        bill_data.append([
            Paragraph(str(i), s_cell),
            Paragraph(str(b.get("bill_date", "")), s_cell),
            Paragraph(str(b.get("bill_no", "")), s_cell),
            Paragraph(str(b.get("customer_name", "")), s_cell_left),
            Paragraph(str(b.get("vehicle_no", "")), s_cell),
            Paragraph(f"{b.get('gross_weight', 0):.2f}", s_cell),
            Paragraph(f"{b.get('tare_weight', 0):.2f}", s_cell),
            Paragraph(f"{b.get('net_weight', 0):.2f}", s_cell),
            Paragraph(f"{b.get('amount', 0):.2f}", s_cell),
        ])

    # Total row for bills
    bill_data.append([Paragraph("", s_cell)] * 7 + [
        Paragraph("<b>Total Amount</b>",
                  ParagraphStyle("TotalLabel", fontName=FONT_NAME, fontSize=8,
                                 alignment=TA_CENTER, textColor=NAVY)),
        Paragraph(f"<b>{total_bills:,.2f}</b>",
                  ParagraphStyle("TotalVal", fontName=FONT_NAME, fontSize=8,
                                 alignment=TA_CENTER, textColor=NAVY)),
    ])

    bill_tbl = Table(bill_data, colWidths=bill_cw, repeatRows=1)
    bill_style = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, -1), FONT_NAME),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 1), (-1, -2), 0.3, BORDER),
        ("LINEABOVE", (0, -1), (-1, -1), 0.8, NAVY),
        ("BACKGROUND", (0, -1), (-1, -1), LIGHT_NAVY),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
    ]
    bill_tbl.setStyle(TableStyle(bill_style))
    elements.append(bill_tbl)
    elements.append(Spacer(1, 3*mm))

    # --- Receipts Table ---
    if receipts:
        elements.append(Paragraph("<b>Receipts / Payments Received</b>", s_section))

        recv_hdr = ["#", "Date", "Receipt No", "Customer", "Mode", "Reference", "Amount"]
        recv_cw = [6*mm, 18*mm, 26*mm, 40*mm, 18*mm, 30*mm, 20*mm]

        recv_data = [[Paragraph(h, s_hdr) for h in recv_hdr]]
        for i, r in enumerate(receipts, 1):
            recv_data.append([
                Paragraph(str(i), s_cell),
                Paragraph(str(r.get("receipt_date", "")), s_cell),
                Paragraph(str(r.get("receipt_no", "")), s_cell),
                Paragraph(str(r.get("customer_name", "")), s_cell_left),
                Paragraph(str(r.get("mode", "")), s_cell),
                Paragraph(str(r.get("reference_no", "")), s_cell),
                Paragraph(f"{r.get('amount', 0):.2f}", s_cell),
            ])

        recv_data.append([Paragraph("", s_cell)] * 5 + [
            Paragraph("<b>Total Received</b>",
                      ParagraphStyle("RTotalLabel", fontName=FONT_NAME, fontSize=8,
                                     alignment=TA_CENTER, textColor=NAVY)),
            Paragraph(f"<b>{total_receipts:,.2f}</b>",
                      ParagraphStyle("RTotalVal", fontName=FONT_NAME, fontSize=8,
                                     alignment=TA_CENTER, textColor=NAVY)),
        ])

        recv_tbl = Table(recv_data, colWidths=recv_cw, repeatRows=1)
        recv_style = [
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, -1), FONT_NAME),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 1), (-1, -2), 0.3, BORDER),
            ("LINEABOVE", (0, -1), (-1, -1), 0.8, NAVY),
            ("BACKGROUND", (0, -1), (-1, -1), LIGHT_NAVY),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ]
        recv_tbl.setStyle(TableStyle(recv_style))
        elements.append(recv_tbl)
        elements.append(Spacer(1, 2*mm))

    # --- Closing Balance ---
    elements.append(Paragraph(
        f"<b>Closing Balance (Pending):  Rs. {balance:,.2f}</b>",
        s_balance,
    ))
    elements.append(Spacer(1, 2*mm))

    # --- Bank Details ---
    bank_parts = []
    if company.get("bank_name"):
        bank_parts.append(f"Bank: {company['bank_name']}")
    if company.get("bank_account"):
        bank_parts.append(f"A/C: {company['bank_account']}")
    if company.get("bank_ifsc"):
        bank_parts.append(f"IFSC: {company['bank_ifsc']}")
    if bank_parts:
        bank_text = "&nbsp;&nbsp;|&nbsp;&nbsp;".join(bank_parts)
        bk_tbl = Table(
            [[Paragraph("<b>Bank Details:</b>",
                         ParagraphStyle("BKL", fontName=FONT_NAME, fontSize=8,
                                        textColor=colors.HexColor("#757575"))),
              Paragraph(bank_text,
                         ParagraphStyle("BKV", fontName=FONT_NAME, fontSize=8,
                                        textColor=colors.black))]],
            colWidths=[26*mm, 146*mm],
        )
        bk_tbl.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ]))
        elements.append(bk_tbl)
        elements.append(Spacer(1, 2*mm))

    # --- Footer ---
    elements.append(HRFlowable(width="100%", thickness=0.3, color=BORDER, spaceAfter=2*mm))
    elements.append(Paragraph(
        "This is a computer-generated report.",
        ParagraphStyle("Footer", fontName=FONT_NAME, fontSize=7,
                       alignment=TA_CENTER, textColor=colors.HexColor("#757575")),
    ))

    doc.build(elements)
    return output_path
